from flask import Flask, jsonify
from flask import request
from flask import send_file
import openpyxl
from werkzeug.utils import secure_filename
app = Flask(__name__)

folder_dir="files"
import os
if not os.path.exists(folder_dir):
    os.makedirs(folder_dir)

def process(file_name):
    wb = openpyxl.load_workbook(file_name)
    sheet=wb.get_sheet_by_name('Table 1')
    wb.active
    wb.create_sheet('Extracted1')
    sheet1=wb.get_sheet_by_name('Extracted1')

    max_row=sheet.max_row
    TS1=sheet.cell(row=(max_row-5),column=1).value
    print(TS1)


    college=sheet.cell(row=1,column=1).value
    exam=sheet.cell(row=4,column=1).value
    s1=sheet.cell(row=6,column=4).value
    s2=sheet.cell(row=6,column=7).value
    s3=sheet.cell(row=6,column=10).value
    s4=sheet.cell(row=6,column=13).value
    s5=sheet.cell(row=6,column=16).value
    s6=sheet.cell(row=6,column=19).value
    s7=sheet.cell(row=6,column=22).value
    s8=sheet.cell(row=6,column=25).value
    s9=sheet.cell(row=6,column=28).value

    sheet1["A1"]="Name"
    sheet1.merge_cells('B1:D1')
    sheet1["B1"]=s1
    sheet1.merge_cells('E1:G1')
    sheet1["E1"]=s2
    sheet1.merge_cells('H1:J1')
    sheet1["H1"]=s3
    sheet1.merge_cells('K1:M1')
    sheet1["K1"]=s4
    sheet1.merge_cells('N1:P1')
    sheet1["N1"]=s5
    sheet1.merge_cells('Q1:S1')
    sheet1["Q1"]=s6
    sheet1.merge_cells('T1:V1')
    sheet1["T1"]=s7
    sheet1.merge_cells('W1:Y1')
    sheet1["W1"]=s8
    sheet1.merge_cells('Z1:AB1')
    sheet1["Z1"]=s9
    sheet1['AC1']="Result"
    sheet1['AD1']="GPA"


    def copyRange(startCol, startRow, endCol, endRow, sheet):
        rangeSelected = []
        #Loops through selected Rows
        for i in range(startRow,endRow + 1,1):
            #Appends the row to a RowSelected list
            rowSelected = []
            for j in range(startCol,endCol+1,1):
                rowSelected.append(sheet.cell(row = i, column = j).value)
            #Adds the RowSelected List and nests inside the rangeSelected
            rangeSelected.append(rowSelected)
     
        return rangeSelected
             
     
    #Paste range
    #Paste data from copyRange into template sheet
    def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
        countRow = 0
        for i in range(startRow,endRow+1,1):
            countCol = 0
            for j in range(startCol,endCol+1,1):
                
                sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
                countCol += 1
            countRow += 1



    def createData():
        print("Processing...")
        selectedRange = copyRange(4,8,30,457,sheet) #Change the 4 number values
        pastingRange = pasteRange(2,2,28,451,sheet1,selectedRange) #Change the 4 number values
        
    createData()




    print("Processing...")
    selectedRange = copyRange(2,8,2,457,sheet) #Change the 4 number values
    pastingRange = pasteRange(1,2,1,451,sheet1,selectedRange) #Change the 4 number values
    print("Range copied and pasted!")


    j=4
    for i in range(1,(TS1+1)):
        sheet1.delete_rows(j)
        sheet1.delete_rows(j)
        sheet1.delete_rows(j)
        j=j+1
        j=j+1




    import re 
    j=2  
    for i in range(1,(TS1+1)): 
       
       test_str = sheet1.cell(row=j,column=1) 

       res= ''.join(filter(lambda k: k.isdigit(),test_str.value))
       s=j+1
       (sheet1.cell(row=s,column=1).value)=res
       j=j+2




    j=8 
    k=2
    for i in range(1,(TS1+1)): 

        s = sheet.cell(row=j,column=31)
        a=re.split('\s+', s.value)
        (sheet1.cell(row=k,column=29).value)=a[1]
        (sheet1.cell(row=k,column=30).value)=a[4]
        j=j+5
        k=k+2
        


    # COUNTING KT STUDENTS 
    flag=0
    value=0
    for i in range(2,(2*(TS1+1))):
        if (flag==0):
            for j in range(2,29):
                if(flag==0):
                    temp1=sheet1.cell(row=i,column=j).value
                    temp=str(temp1)
                    for char in temp:
                        if char is '+':
                            value=i
                            flag=1
                            break
                        else:
                            continue
                
                    j=j+1
                else:
                    break
            i=i+1
        else:
            break
    print(value)
    kt=TS1-(value/2)+1

    kt=int(kt)

    print(kt)
    TS=TS1-kt

    j=156
    for i in range(1,(2*kt)+1):
        sheet1.delete_rows(j)
        

    wb.save("result_output.xlsx")

    pcount=0
    fcount=0
    k=2
    for i in range(1,(TS+1)):
        if((sheet1.cell(row=k,column=29).value)=="P"):
            pcount=pcount+1
        if((sheet1.cell(row=k,column=29).value)=="F"):
            fcount=fcount+1
        k=k+2
    print(pcount)
    print(fcount)

    import array as arr
    a=arr.array('i',[0,0,0,0,0,0,0,0,0,0])
    count=0
    y=1
    k=3
    l=4
    while(l<29):
        for i in range(1,(TS+1)):
        
            if((sheet1.cell(row=k,column=l).value)=="--"):
                count=count+1
            k=k+2
        
        a[y]=count
        y=y+1
        count=0
        k=3
        l=l+3
    for i in range(1,10):
        print(a[i])
    a1=a[1]
    a2=a[2]
    a3=a[3]
    a4=a[4]
    a5=a[5]
    a6=a[6]
    a7=a[7]
    a8=a[8]
    a9=a[9]



    wb.save("result_output.xlsx")

    import pandas as pd

    xl = pd.ExcelFile("result_output.xlsx")
    df = xl.parse("Extracted1")
    df = df.sort_values(["GPA"],ascending=False)

    writer = pd.ExcelWriter("cal.xlsx")
    df.to_excel(writer,sheet_name='sheet1',columns=["Name","Result","GPA"],index=False)
    writer.save()

    wb2 = openpyxl.load_workbook('cal.xlsx')
    sheet5=wb2.get_sheet_by_name('sheet1')
    wb2.active


    distinction=0
    first=0
    second=0
    f=0
    k=2

    for i in range(1,(TS+1)):
        
        a=sheet5.cell(row=k,column=3).value
        
        if(a=='--'):
            f=f+1
            break
        else:
            a=float(sheet5.cell(row=k,column=3).value)
            if(a>=7.75):
                distinction=distinction+1
                k=k+1
            elif(a>=6.75 ):
                first=first+1
                k=k+1
            elif(a<6.75 ):
                second=second+1
                k=k+1
            else:
                f=f+1
            
        
    print(distinction)
    print(first)
    print(second)

    wb.create_sheet('Result Analysis')
    sheet3=wb.get_sheet_by_name('Result Analysis')
    wb.active

    sheet3.merge_cells('B1:H1')
    sheet3.merge_cells('A2:K2')
    sheet3.merge_cells('B3:D3')
    sheet3.merge_cells('B4:D4')
    sheet3['B1']=college
    sheet3['A2']=exam
    sheet3['B4']="Result Analysis "

    sheet3['B6']="Class"
    sheet3['B7']="No. of Student Appeared"
    sheet3['B8']="No. of Student passed"
    sheet3['B9']="No. of Student failed"
    sheet3['B10']="No. of Student having SGPA 7.75 and above (Distinction)"
    sheet3['B11']="No. of Student having SGPA 6.75 and above (First)"
    sheet3['B12']="No. of Student having SGPA below 6.75 (Second )"
    sheet3["B13"]="% of passing"
    sheet3['C6']="S.E.I.T"
    sheet3['C7']=TS
    sheet3['C8']=pcount
    sheet3['C9']=fcount
    sheet3['C10']=distinction
    sheet3['C11']=first
    sheet3['C12']=second
    sheet3["C13"]=(pcount/TS)*100


    sheet3['B15']="Subject Wise Analysis"
    sheet3['A17']="SR No"
    sheet3['B17']="Subject"
    sheet3['C17']="No. of candidates appeared"
    sheet3['D17']="No. of candidates passed"
    sheet3['E17']="% of passing in the subject"
    sheet3['A18']="1"
    sheet3['A19']="2"
    sheet3['A20']="3"
    sheet3['A21']="4"
    sheet3['A22']="5"
    sheet3['A23']="6"
    sheet3['A24']="7"
    sheet3['A25']="8"
    sheet3['A26']="9"
    sheet3['B18']=s1
    sheet3['B19']=s2
    sheet3['B20']=s3
    sheet3['B21']=s4
    sheet3['B22']=s5
    sheet3['B23']=s6
    sheet3['B24']= s7
    sheet3['B25']=s8
    sheet3['B26']=s9       

    sheet3["C18"]=TS
    sheet3['C19']=TS
    sheet3["C20"]=TS
    sheet3["C21"]=TS
    sheet3["C22"]=TS
    sheet3["C23"]=TS
    sheet3["C24"]=TS
    sheet3["C25"]=TS
    sheet3["C26"]=TS
    sheet3["D18"]=(TS)-a1
    sheet3['D19']=TS-a2
    sheet3["D20"]=TS-a3
    sheet3["D21"]=TS-a4
    sheet3["D22"]=TS-a5
    sheet3["D23"]=TS-a6
    sheet3["D24"]=TS-a7
    sheet3["D25"]=TS-a8
    sheet3["D26"]=TS-a9
    sheet3["E18"]=((TS-a1)/TS1)*100
    sheet3['E19']=((TS-a2)/TS1)*100
    sheet3["E20"]=((TS-a3)/TS1)*100
    sheet3["E21"]=((TS-a4)/TS1)*100
    sheet3["E22"]=((TS-a5)/TS1)*100
    sheet3["E23"]=((TS-a6)/TS1)*100
    sheet3["E24"]=((TS-a7)/TS1)*100
    sheet3["E25"]=((TS-a8)/TS1)*100
    sheet3["E26"]=((TS-a9)/TS1)*100

    wb.save("result_output.xlsx")

    wb.create_sheet('Topper')
    sheet2=wb.get_sheet_by_name('Topper')

    sheet2["A6"]="Student Name  Seat No"
    sheet2.merge_cells('A1:H1')
    sheet2.merge_cells('A2:K2')
    sheet2.merge_cells('B3:E3')
    sheet2.merge_cells('A4:F4')
    sheet2["A1"]=college
    sheet2["A2"]=exam
    sheet2["A4"]="Heartiest Congratulations To All The Toppers"

    sheet2["B6"]="College Rank"
    sheet2["C6"]="SGPA"



    p=sheet5.cell(row=2,column=3).value

    r=1
    sheet2.cell(row=7,column=1).value=sheet5.cell(row=2,column=1).value
    sheet2.cell(row=7,column=2).value=1
    sheet2.cell(row=7,column=3).value=sheet5.cell(row=2,column=3).value
    k=3

    for i in range(1,(TS1+1)):
        while(r<5):
            sp=sheet5.cell(row=k,column=3).value
            if(sp==p):
                sheet2.cell(row=k+5,column=1).value=sheet5.cell(row=k,column=1).value
                sheet2.cell(row=k+5,column=2).value=r
                sheet2.cell(row=k+5,column=3).value=sheet5.cell(row=k,column=3).value
            else:
                r=r+1
                sheet2.cell(row=k+5,column=1).value=sheet5.cell(row=k,column=1).value
                sheet2.cell(row=k+5,column=2).value=r
                sheet2.cell(row=k+5,column=3).value=sheet5.cell(row=k,column=3).value
                p=sp
            k=k+1
                
                
    wb.save("result_output.xlsx")

    wb.remove_sheet(sheet1)

    wb.save("result_output.xlsx")
    return send_file('result_output.xlsx', attachment_filename='result_output.xlsx')


@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        #image=request.args.get('image')
        file=request.files['myFile']
        filename = secure_filename(file.filename)
        upload_file=folder_dir+"/"+filename
        file.save(upload_file)
        #with open(filename, "rb") as image_file:
         #   encoded_string = base64.b64encode(image_file.read())
        print("excel recieved")
        print(upload_file)
        return process(upload_file)
    
    return '''
    <html>
<head>
<meta name="viewport" content="width=device-width, initial-scale=1">
<style>
html {
  scroll-behavior: smooth;
}

body {
  font-family: Arial, Helvetica, sans-serif;
}

.header {
  padding: 40px;
  background: white;
  color: #b30000;
  font-size: 30px;
}


.navbar {
  overflow: hidden;
  background-color: #333;
}

.navbar a {
  float: left;
  font-size: 16px;
  color: white;
  text-align: center;
  padding: 14px 16px;
  text-decoration: none;
}

.navbar a1 {
  float: left;
  font-size: 16px;
  color: white;
  text-align: center;
  padding: 14px 16px;
  text-decoration: none;
}

.dropdown {
  float: left;
  overflow: hidden;
}

.dropdown .dropbtn {
  font-size: 16px;  
  border: none;
  outline: none;
  color: white;
  padding: 14px 16px;
  background-color: inherit;
  font-family: inherit;
  margin: 0;
}

.navbar a:hover, .dropdown:hover .dropbtn {
  background-color: white;
  color: black;
}

.dropdown-content {
  display: none;
  position: absolute;
  background-color: #f9f9f9;
  min-width: 160px;
}

.dropdown-content a {
  float: none;
  color: black;
  padding: 12px 16px;
  text-decoration: none;
  display: block;
  text-align: left;
}

.dropdown-content a:hover {
  background-color: #ddd;
}

.dropdown:hover .dropdown-content {
  display: block;
}

#section2 {
  height: 600px;
  background-color: yellow;
</style>
</head>
<body>
<div class="header">
	<h1 align="center">K.J. SOMAIYA INSTITUTE OF ENGINEERING AND INFORMATION TECHNOLOGY</h1>
	
</div>
<div>
	<h2><center>RESULT ANALYSIS</center></h2>
</div>

<div>
 <form method=post enctype=multipart/form-data>
   Select a file: <input type="file" name="myFile"><br><br>
   <input type=submit value=Upload>
</form>



</body>
</html>

    '''

if __name__ == '__main__':
    #app.run(host='localhost', port=5000) #run in local
    app.run()


