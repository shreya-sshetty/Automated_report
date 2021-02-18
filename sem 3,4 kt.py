#!/usr/bin/env python
# coding: utf-8

# In[22]:


import openpyxl
wb = openpyxl.load_workbook('result.xlsx')
sheet=wb.get_sheet_by_name('Table 1')
wb.active
wb.create_sheet('Extracted1')
sheet1=wb.get_sheet_by_name('Extracted1')

max_row=sheet.max_row
TS1=sheet.cell(row=(max_row-5),column=1).value
print(TS1)


# In[23]:


college=sheet.cell(row=1,column=1).value
exam=sheet.cell(row=4,column=1).value
s1=sheet.cell(row=6,column=4).value
s2=sheet.cell(row=6,column=7).value
s3=sheet.cell(row=6,column=10).value
s4=sheet.cell(row=6,column=13).value
s5=sheet.cell(row=6,column=16).value
s6=sheet.cell(row=6,column=19).value
s7=sheet.cell(row=6,column=22).value
s8=sheet.cell(row=6,column=25).value
s9=sheet.cell(row=6,column=28).value


sheet1["A1"]="Name"
sheet1.merge_cells('B1:D1')
sheet1["B1"]=s1
sheet1.merge_cells('E1:G1')
sheet1["E1"]=s2
sheet1.merge_cells('H1:J1')
sheet1["H1"]=s3
sheet1.merge_cells('K1:M1')
sheet1["K1"]=s4
sheet1.merge_cells('N1:P1')
sheet1["N1"]=s5
sheet1.merge_cells('Q1:S1')
sheet1["Q1"]=s6
sheet1.merge_cells('T1:V1')
sheet1["T1"]=s7
sheet1.merge_cells('W1:Y1')
sheet1["W1"]=s8
sheet1.merge_cells('Z1:AB1')
sheet1["Z1"]=s9
sheet1['AC1']="Result"
sheet1['AD1']="GPA"



def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
 
    return rangeSelected
         
 
#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1



def createData():
    print("Processing...")
    selectedRange = copyRange(4,8,30,457,sheet) #Change the 4 number values
    pastingRange = pasteRange(2,2,28,451,sheet1,selectedRange) #Change the 4 number values
    
createData()




print("Processing...")
selectedRange = copyRange(2,8,2,457,sheet) #Change the 4 number values
pastingRange = pasteRange(1,2,1,451,sheet1,selectedRange) #Change the 4 number values
print("Range copied and pasted!")

j=4
for i in range(1,(TS1+1)):
    sheet1.delete_rows(j)
    sheet1.delete_rows(j)
    sheet1.delete_rows(j)
    j=j+1
    j=j+1
    
import re 
j=2  
for i in range(1,(TS1+1)): 
   
   test_str = sheet1.cell(row=j,column=1) 

   res= ''.join(filter(lambda k: k.isdigit(),test_str.value))
   s=j+1
   (sheet1.cell(row=s,column=1).value)=res
   j=j+2

j=8 
k=2
for i in range(1,(TS1+1)): 

    s = sheet.cell(row=j,column=31)
    a=re.split('\s+', s.value)
    (sheet1.cell(row=k,column=29).value)=a[1]
    (sheet1.cell(row=k,column=30).value)=a[4]
    j=j+5
    k=k+2

wb.save("result.xlsx")


# In[24]:


TS=TS1
pcount=0
fcount=0
k=2
for i in range(1,(TS+1)):
    if((sheet1.cell(row=k,column=29).value)=="P"):
        pcount=pcount+1
    if((sheet1.cell(row=k,column=29).value)=="F"):
        fcount=fcount+1
    k=k+2
print(pcount)
print(fcount)

import array as arr
a=arr.array('i',[0,0,0,0,0,0,0,0,0,0,0])
count=0
y=1
k=3
l=4
while(l<29):
    for i in range(1,(TS+1)):
    
        if((sheet1.cell(row=k,column=l).value)=="--"):
            count=count+1
        k=k+2
    
    a[y]=count
    y=y+1
    count=0
    k=3
    l=l+3
for i in range(1,11):
    print(a[i])
a1=a[1]
a2=a[2]
a3=a[3]
a4=a[4]
a5=a[5]
a6=a[6]
a7=a[7]
a8=a[8]
a9=a[9]
a10=a[10]



wb.save("result.xlsx")


# In[25]:


import pandas as pd

xl = pd.ExcelFile("result.xlsx")
df = xl.parse("Extracted1")
df = df.sort_values(["GPA"],ascending=False)

writer = pd.ExcelWriter("output.xlsx")
df.to_excel(writer,sheet_name='sheet1',columns=["Name","Result","GPA"],index=False)
writer.save()

wb2 = openpyxl.load_workbook('output.xlsx')
sheet5=wb2.get_sheet_by_name('sheet1')
wb2.active






import re 
j=2  
for i in range(1,(TS+1)): 
   
   test_str = sheet5.cell(row=j,column=1) 

   res= ''.join(filter(lambda k: k.isdigit(),test_str.value))
   (sheet5.cell(row=j,column=4).value)=res
   j=j+1

for k in range(2,(TS+1)):
    t1=sheet5.cell(row=k,column=1).value
    t2=sheet5.cell(row=k,column=4).value
    l1 = t1.lower().split()
    l2 = t2.lower().split()
    h1 = ""
    h2 = ""
    for i in l1:
      if i not in l2:
        h1 = h1 + " " + i 
    for j in l2:
      if j not in l1:
        h2 = h2 + " " + j 

    new = h1 + " " + h2
    sheet5.cell(row=k,column=1).value=new
wb2.save("output.xlsx")    














wb.save("result.xlsx")    
distinction=0
first=0
second=0
f=0
k=2

for i in range(1,(TS+1)):
    
    a=sheet5.cell(row=k,column=3).value
    
    if(a=='--'):
        f=f+1
        break
    else:
        a=float(sheet5.cell(row=k,column=3).value)
        if(a>=7.75):
            distinction=distinction+1
            k=k+1
        elif(a>=6.75 ):
            first=first+1
            k=k+1
        elif(a<6.75 ):
            second=second+1
            k=k+1
        else:
            f=f+1
        
    
print(distinction)
print(first)
print(second)


# In[26]:


from openpyxl.styles import colors
from openpyxl.styles import Font,Color
ft=Font

wb.create_sheet('Result Analysis')
sheet3=wb.get_sheet_by_name('Result Analysis')
wb.active

sheet3.merge_cells('C1:H1')
sheet3.merge_cells('D2:G2')
sheet3.merge_cells('B3:I3')
sheet3.merge_cells('D4:G4')
sheet3.merge_cells('C5:H5')

sheet3['C1']=college
sheet3["D2"]="NAAC Accredited Institute with “A” Grade"
sheet3["B3"]="NBA Accredited 3 Programs (Computer Engineering, Electronics &Telecommunication"
sheet3["D4"]="Engineering & Electronics Engineering)"
sheet3["C5"]="Permanently Affiliated to University of Mumbai"
sheet3["C6"]= "DEPARTMENT OF INFORMATION TECHNOLOGY"
sheet3['B10']="Result Analysis "
sheet3['A8']=exam
sheet3.merge_cells('A8:J8')

sheet3.merge_cells('B18:F18')
sheet3.merge_cells('B19:F19')
sheet3.merge_cells('B13:F13')
sheet3.merge_cells('B14:F14')
sheet3.merge_cells('B15:F15')
sheet3.merge_cells('B16:F16')
sheet3.merge_cells('B17:F17')




sheet3['B12']="Class"
sheet3['B13']="No. of Student Appeared"
sheet3['B14']="No. of Student passed"
sheet3['B15']="No. of Student failed"
sheet3['B16']="No. of Student having SGPA 7.75 and above (Distinction)"
sheet3['B17']="No. of Student having SGPA 6.75 and above (First)"
sheet3['B18']="No. of Student having SGPA below 6.75 (Second )"
sheet3["B19"]="% of passing"
sheet3['G12']="S.E.I.T"
sheet3['G13']=TS
sheet3['G14']=pcount
sheet3['G15']=fcount
sheet3['G16']=distinction
sheet3['G17']=first
sheet3['G18']=second-fcount
sheet3["G19"]=(pcount/TS)*100


sheet3.merge_cells('B30:E30')
sheet3.merge_cells('B31:E31')
sheet3.merge_cells('B32:E32')
sheet3.merge_cells('B24:E24')
sheet3.merge_cells('B25:E25')
sheet3.merge_cells('B26:E26')
sheet3.merge_cells('B27:E27')
sheet3.merge_cells('B28:E28')
sheet3.merge_cells('B29:E29')
sheet3['B21']="Subject Wise Analysis"
sheet3['A23']="SR No"
sheet3['B23']="Subject"
sheet3['F23']="Appeared"
sheet3['G23']="Passed"
sheet3['H23']="% of passing subject"
sheet3['A24']="1"
sheet3['A25']="2"
sheet3['A26']="3"
sheet3['A27']="4"
sheet3['A28']="5"
sheet3['A29']="6"
sheet3['A30']="7"
sheet3['A31']="8"
sheet3['A32']="9"

sheet3['B24']=s1
sheet3['B25']=s2
sheet3['B26']=s3
sheet3['B27']=s4
sheet3['B28']=s5
sheet3['B29']=s6
sheet3['B30']= s7
sheet3['B31']=s8
sheet3['B32']=s9       
     

sheet3["F24"]=TS
sheet3['F25']=TS
sheet3["F26"]=TS
sheet3["F27"]=TS
sheet3["F28"]=TS
sheet3["F29"]=TS
sheet3["F30"]=TS
sheet3["F31"]=TS
sheet3["F32"]=TS

sheet3["G24"]=(TS)-a1
sheet3['G25']=TS-a2
sheet3["G26"]=TS-a3
sheet3["G27"]=TS-a4
sheet3["G28"]=TS-a5
sheet3["G29"]=TS-a6
sheet3["G30"]=TS-a7
sheet3["G31"]=TS-a8
sheet3["G32"]=TS-a9

sheet3["H24"]=((TS-a1)/TS1)*100
sheet3['H25']=((TS-a2)/TS1)*100
sheet3["H26"]=((TS-a3)/TS1)*100
sheet3["H27"]=((TS-a4)/TS1)*100
sheet3["H28"]=((TS-a5)/TS1)*100
sheet3["H29"]=((TS-a6)/TS1)*100
sheet3["H30"]=((TS-a7)/TS1)*100
sheet3["H31"]=((TS-a8)/TS1)*100
sheet3["H32"]=((TS-a9)/TS1)*100


wb.save("result.xlsx")


# In[ ]:


import openpyxl

wb = openpyxl.load_workbook('result.xlsx')
sheet3=wb.get_sheet_by_name('Result Analysis')
wb.active

img = openpyxl.drawing.image.Image('python.png')
img.anchor = 'A1'
img.width = 60
img.height = 40
sheet3.add_image(img)
img = openpyxl.drawing.image.Image('python2.png')
img.anchor = 'J1'
img.width = 60
img.height = 40
sheet3.add_image(img)
wb.save('result.xlsx')


wb.save('result.xlsx')


# In[ ]:


sheet3["C1"].font = Font(bold=True)
sheet3["C6"].font = Font(bold=True)
sheet3["B10"].font = Font(bold=True)
sheet3["B21"].font = Font(bold=True)

wb.save("result.xlsx")


# In[ ]:


import matplotlib.pyplot as plt

import openpyxl

# Pie chart
labels = ['Failed', '7.5 and above', '6.5 and above', 'below 6.5']
sizes = [sheet3['G15'].value, sheet3['G16'].value, sheet3['G17'].value, sheet3['G18'].value]
# only "explode" the 2nd slice (i.e. 'Hogs')
explode = (0, 0.1, 0, 0)  
fig1, ax1 = plt.subplots()
ax1.pie(sizes, explode=explode, labels=labels, autopct='%1.1f%%',shadow=True, startangle=90)
# Equal aspect ratio ensures that pie is drawn as a circle
ax1.axis('equal')  
plt.tight_layout()

# Your plot generation code here...
plt.savefig("myplot.png", dpi = 150)

wb = openpyxl.load_workbook('result.xlsx')
sheet3=wb.get_sheet_by_name('Result Analysis')


img = openpyxl.drawing.image.Image('myplot.png')
img.anchor = 'B34'
img.width = 400
img.height = 300
sheet3.add_image(img)

wb.save('result.xlsx')


# In[ ]:



from win32com import client
import win32api

def exceltopdf(doc):
    excel = client.DispatchEx("Excel.Application")
    excel.Visible = 0

    wb = excel.Workbooks.Open('C:/Users/Radhika Choudhary/result.xlsx')
    ws = wb.Worksheets[1]

    try:
        wb.SaveAs('C:/Users/Radhika Choudhary/result.pdf', FileFormat=57)
    except Exception:
        print ("Failed to convert")
   
    finally:
        wb.Close()
        excel.Quit()


# In[ ]:




